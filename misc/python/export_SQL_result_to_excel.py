'''
Read all the SQL files within the folder
Execute them against the database 
Save the result into one Excel
'''

import pandas as pd
import pyodbc
import os
import warnings
import pymssql
import openpyxl

warnings.filterwarnings('ignore') # do not show Python package internal warnings


# server = 'server:port'
# db = 'dbName'
# username = 'usr'
# password = 'pwd'
# scriptsFolder = r'path_to_sql_folder'
# outputExcel = r'path_to_output_excel_file'


#cnxn = pyodbc.connect(f'Driver=SQL Server;Server={server};Database={db};Trusted_Connection=yes;')
connection = pymssql.connect(host=server, database=db, user=username, password=pwd)

allFiles = os.listdir(scriptsFolder)
#print(arr)
sheet_cnt = 0

# if os.path.isfile(outputExcel):
#     writer = pd.ExcelWriter(outputExcel, mode='a', if_sheet_exists = 'replace')
# else:
#     writer = pd.ExcelWriter(outputExcel)

for file in allFiles:
    realPath = scriptsFolder + '/' + file
    if realPath[-4:].upper() == '.SQL': # only take the SQL file
        sheet_cnt += 1
        print(f'Working on {sheet_cnt} of total {len(allFiles)} files: ' + file)
        
        with open(realPath, 'r') as f:
            sql = "" # otherwise, pd.read_sql will return the first sql result. It returns None for the ones creating temp table or the ones with "Warning: Null value is eliminated by an aggregate or other SET operation."
            for _ in f.readlines():
                sql += _ 
        
        cursor = connection.cursor()
        cursor.execute(sql)
        book = openpyxl.load_workbook(outputExcel)
        print('names:' , book.sheetnames)
        sheet = book.get_sheet_by_name("Sheet_1")
        r = 1 
        c = 1 
        for col_name in cursor.description:
            sheet.cell(row=r, column=c).value = col_name[0]
            c += 1
        r += 1
        
        for currentRow in cursor:
            arr = list(currentRow)
            c = 1
            for content in arr:
               sheet.cell(row=r, column=c).value = content
               c+=1
            r += 1 

book.save(outputExcel)
        # df = pd.DataFrame.from_records(cursor.fetchall(),columns = [desc[0] for desc in cursor.description])
        # df.to_excel(writer, sheet_name=f'Sheet_{sheet_cnt}', index=False,)
        
        # ws = writer.sheets[f'Sheet_{sheet_cnt}']
        # ws.hide()

# writer.save()




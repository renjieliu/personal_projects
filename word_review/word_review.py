import random
import datetime as dt
import os
from openpyxl import load_workbook
import dotenv

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

dotenv.load_dotenv('.env')

excel_path = r"C:\Users\rliu7.GLOBAL\Desktop\Word list.xlsx"

# Load the workbook and select a sheet
wb = load_workbook(excel_path)

sheet = wb['English'] #to select a specific sheet

# # Access data from specific cells
# value = sheet['A1'].value  # Get the value from cell A1
# #print(value)

excel_dump  = [] 
# Loop through rows and columns to read multiple values
for row in sheet.iter_rows() : # min_row=2, max_col=3, max_row=5):
    excel_dump.append([])
    for cell in row:
        excel_dump[-1].append(cell.value)

word_array = []

for rows in excel_dump:
    if rows[2] != None: # the word itself is in the third column, and the added date is on the fourth column
        txt =  'Word: ' + rows[2]  + ' -->  ' + ' added on: '  + str(rows[3]).split(' ')[0]
        word_array.append(txt)
        #print('Word: ', rows[2], ' , ' ,'Added on:', str(rows[3]).split(' ')[0] )


random.shuffle(word_array)
pick_count = 10
pick_count = pick_count if pick_count <= len(word_array) else len(word_array)

today_pick = word_array[:pick_count]

# print(today_pick)

######### to send out a mail ######### 



# create a message
msg = MIMEMultipart()
mailFrom = os.getenv('mailFrom')
mailTo=os.getenv('mailTo')
gmailToken = os.getenv('gmailToken')

msg['From'] = mailFrom
msg['To'] = mailTo
msg['Subject'] = r'''Today's words'''
body = '\r\n'.join(today_pick)

msg.attach(MIMEText(body, 'plain'))

# set up the SMTP server
server = smtplib.SMTP('smtp.gmail.com', 587)
server.starttls()
server.login(mailFrom, gmailToken)

# send the message
server.sendmail(mailFrom, mailTo, msg.as_string())
print('Email sent!')

# close the connection
server.quit()

